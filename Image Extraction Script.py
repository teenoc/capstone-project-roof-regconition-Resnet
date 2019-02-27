import pyautogui
from pywinauto.application import Application
import time
import os
import subprocess
import pyperclip
import PIL
from PIL import ImageGrab
from openpyxl.reader.excel import load_workbook
from platform import python_version

#The key to google staic map api
key = "AIzaSyCkUOdZ5y7hMm0yrcCQoCvLwzdM6M8s5qk"
#This is where your excel file with labelled addresses is located, please follow the format as in the given 'labelled addresses.xlsx' file.
exlpath="C:/Users/Teeno/Desktop/result/new_address.xlsx"
#This is where you want to download your images to
image =r"C:/Users/Teeno/Desktop/result/predict/"
#the path to google chrome
chromepath=r"C:/Program Files (x86)/Google/Chrome/Application/chrome.exe"
#two websites for downloading the images
chromeurl_coor="https://www.latlong.net"
chromeurl_image="https://jsfiddle.net/api/post/library/pure/"

#the name of the download images
path_45 =image +'45_%d_%d.jpg'
path_top =image +'top_%d.jpg'
#to count the rows in the excel file
row_count = load_workbook(exlpath).worksheets[0].max_row
maxrow = row_count+1
x=[]
y=[]

#the content to be used on the website to query the images
HTML = '''<div id="map"></div>\n<!-- Replace the value of the key parameter with your own API key. -->\n<script async defer\nsrc="https://maps.googleapis.com/maps/api/js?key=%s
&callback=initMap">\n</script>'''%key
CSS = '''#map {\n  height: 100%;\n}\nhtml, body {\n  height: 100%;\n  margin: 0;\n  padding: 0;\n}'''
JavaScript = '''function initMap() {\n  var map = new\ngoogle.maps.Map(document.getElementById('map'), {\n    center: {lat: 36.964, lng: -122.015},\n    zoom: 18,\n    mapTypeId: 'satellite'\n  });\n  map.setTilt(45);\n}'''



####****when used on a new device, please run swapy64bit.exe to make sure properties of the following functions are correct to ensure a smooth execution***#####
####****crash might also be related to sleep time being not sufficient between steps****####

##this is to restore python shell to screen
def shell():
    app = Application().Connect(title=u'*Python 3.6.4 Shell*', class_name='TkTopLevel')
    tktoplevel = app.TkTopLevel
    tktoplevel.Restore()
    tktoplevel.ClickInput()

##minimize python shell
def shellMin():
    app = Application().Connect(title=u'*Python 3.6.4 Shell*', class_name='TkTopLevel')
    tktoplevel = app.TkTopLevel
    tktoplevel.Minimize()

##maximize excel file
def exlMax():
    app = Application().Connect(title=os.path.basename(exlpath) + ' - Excel', class_name='XLMAIN')
    xlmain = app.XLMAIN
    excel = xlmain.EXCEL7
    xlmain.Maximize()
    xlmain.SetFocus()
    
##minimize excel file
def exlMin():
    app = Application().Connect(title=os.path.basename(exlpath) + ' - Excel', class_name='XLMAIN')
    xlmain = app.XLMAIN
    excel = xlmain.EXCEL7
    xlmain.Minimize()

##maximize google chrome window
def chromeMax():
    app = Application().Connect(title=u'Latitude and Longitude Finder on Map Get Coordinates - Google Chrome', class_name='Chrome_WidgetWin_1')
    chromewidgetwin = app.Chrome_WidgetWin_1
    chromewidgetwin.Maximize()
    chromewidgetwin.SetFocus()

    
##maximize google chrome window: Lat and Long website
def chromeMax():
    app = Application().Connect(title=u'Latitude and Longitude Finder on Map Get Coordinates - Google Chrome', class_name='Chrome_WidgetWin_1')
    chromewidgetwin = app.Chrome_WidgetWin_1
    chromewidgetwin.Maximize()
    chromewidgetwin.SetFocus()
    
##minimize google chrome window:Lat and Long website
def chromeMin():
    app = Application().Connect(title=u'Latitude and Longitude Finder on Map Get Coordinates - Google Chrome', class_name='Chrome_WidgetWin_1')
    chromewidgetwin = app.Chrome_WidgetWin_1
    chromewidgetwin.Minimize()
    
##maximize google chrome window:JSFiddle
def JSMax():
    app = Application().Connect(title=u'Edit fiddle - JSFiddle - Google Chrome', class_name='Chrome_WidgetWin_1')
    chromewidgetwin = app.Chrome_WidgetWin_1
    chromewidgetwin.Maximize()
    chromewidgetwin.SetFocus()

##minimize google chrome window:JSFiddle
def JSMin():
    app = Application().Connect(title=u'Edit fiddle - JSFiddle - Google Chrome', class_name='Chrome_WidgetWin_1')
    chromewidgetwin = app.Chrome_WidgetWin_1
    chromewidgetwin = app.Chrome_WidgetWin_1
    chromewidgetwin.Minimize()

#record mouse location for JSFiddle
def recordScript(x,y,z):
    for i in range(0,100):
        question = input("Cursor in place? y/n: ")
        if question == 'n':
            print ('Try again, please move cursor to the specified region')
        elif question == 'y':
            pyautogui.click()
            pyautogui.hotkey('ctrl', 'a')
            x, y = pyautogui.position()
            file = open("script_position.txt","a")
            file.write(str(x))
            file.write("\n")
            file.write(str(y))
            file.write("\n")
            file.close() 
            pyperclip.copy(z)
            pyautogui.hotkey('ctrl', 'v')
            break

#record mouse location and generate image query script on JSFiddle
def configure():
    chromeMax()
    pyautogui.hotkey('ctrl', 'Tab')
    shell()
    for i in range(0,100):
        file = open("script_position.txt","w")
        file.close() 
        print ('Please move cursor over HTML region')
        recordScript(x,y,HTML)
        shell()
        print ('Please move cursor over CSS region')
        recordScript(x,y,CSS)
        shell()
        print ('Please move cursor over JavaScript region')
        recordScript(x,y,JavaScript)
        shell()
        question = input("Generation successful? y/n: ")
        if question == 'y':
            break
    pyautogui.hotkey('alt', 'Tab')
    pyautogui.hotkey('ctrl', 'Tab')
    chromeMin()
    shell()

#generate image query script on JSFiddle
def generateScript():
        chromeMax()
        pyautogui.hotkey('ctrl', 'Tab')
        s1,s2,s3= (0,0,0)
        z1,z2,z3= (0,0,0)
        with open("script_position.txt") as f:
            content = f.readlines()
        s1,z1 =(int(content[0]),int(content[1]))
        s2,z2 =(int(content[2]),int(content[3]))
        s3,z3 =(int(content[4]),int(content[5]))          
        pyautogui.click(s1,z1)
        pyautogui.hotkey('ctrl', 'a')
        pyperclip.copy(HTML)
        pyautogui.hotkey('ctrl', 'v')
        pyautogui.click(s2,z2)
        pyautogui.hotkey('ctrl', 'a')
        pyperclip.copy(CSS)
        pyautogui.hotkey('ctrl', 'v')    
        pyautogui.click(s3,z3)
        pyautogui.hotkey('ctrl', 'a')
        pyperclip.copy(JavaScript)
        pyautogui.hotkey('ctrl', 'v')
        pyautogui.hotkey('ctrl', 'Tab')
        chromeMin()

#record mouse loaction 
def getposition(): 
    for a in range(0,100):
        question = input("Cursor in place? y/n: ")
        if question == 'n':
            print ('Try again, no need to click')
        elif question == 'y':
            x, y = pyautogui.position()
            pyautogui.click()
            file = open("record.txt","a")
            file.write(str(x))
            file.write("\n")
            file.write(str(y))
            file.write("\n")
            file.close() 
            print(x,y)
            break

#record mouse loaction for Lat and Long website
def recordMouse():
    chromeMax()
    for j in range(0,100):
        file = open("record.txt","w")
        file.close()
        shell()
        print("Please move cursor to blank area of cell 'Place Name'.\n(Where it says 'Type a place name')")
        getposition()
        shell()
        print("Please move cursor to blank area of cell 'Latitude'.\n(Where it says 'lat coordinate')")
        getposition()
        shell()
        print("Please move cursor to blank area of cell 'Longitude'.\n(Where it says 'long coordinate')")
        getposition()
        pyautogui.hotkey('ctrl', 'Tab') 
        shell()
        print("Please move cursor to the fourth line of JavaScript'.\n(It looks like'center: {lat: 36.964, lng: -122.015},')")
        getposition()
        shell()
        print("Please move cursor to 'Run' button at top left corner.")
        getposition()
        time.sleep(1)
        shell()
        print("Please move cursor to 'Toggle' button at top right corner of the image.")
        getposition()
        shell()
        print("Please move cursor to 'Rotate' button.\nUsually it's the first button at the bottom right corner.")
        getposition()
        shell()
        print("Please move cursor to 'Change-View' button.\nUsually it's the second button at the bottom right corner.")
        getposition()
        shell()
        print("Please move cursor to 'Zoom-In' button.\nUsually it's the fourth button at the bottom right corner.")
        getposition()
        pyautogui.press('esc')
        pyautogui.hotkey('ctrl', 'Tab')
        shell()
        question = input("All mouse location recorded? y/n: ")
        if question == 'y':
            chromeMin()
            break
        elif question =='n':
            print("Please try again")

#click into excel combobox, please use swapy to configure /combobox = xlamin[u'9']/ for new device              
def combobox(x):
    app = Application().Connect(title=os.path.basename(exlpath) + ' - Excel', class_name='XLMAIN')
    xlmain = app.XLMAIN
    xlmain.Maximize()
    combobox = xlmain[u'10']
    combobox.DoubleClick()
    pyperclip.copy('A%d'%x)
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.press('enter')
    pyautogui.hotkey('ctrl', 'c')
    xlmain.Minimize()

#click into excel combobox, please use swapy to configure /combobox = xlamin[u'9']/ for new device    
def clickin():
    app = Application().Connect(title=os.path.basename(exlpath) + ' - Excel', class_name='XLMAIN')
    xlmain = app.XLMAIN
    xlmain.Maximize()
    combobox = xlmain[u'10']
    combobox.DoubleClick()
    pyautogui.press('enter')

#this will automatically search for coordinates from Lat and Long website in each row of the excel file, and download images from JSFiddle
#configure box to obtain desired image size, currently is set at 500*500 pixels at the center of the image
#configure the first for loop to download images of address of specific row in excel
#configure time.sleep() if necessary when website is responding slowly
def ImageGrab():
    x1,x2,x3,x4,x5,x6,x7,x8,x9= (0,0,0,0,0,0,0,0,0)
    y1,y2,y3,y4,y5,y6,y7,y8,y9= (0,0,0,0,0,0,0,0,0)
    with open("record.txt") as f:
        content = f.readlines()
    x1,y1 =(int(content[0]),int(content[1]))
    x2,y2 =(int(content[2]),int(content[3]))
    x3,y3 =(int(content[4]),int(content[5]))
    x4,y4 =(int(content[6]),int(content[7]))
    x5,y5 =(int(content[8]),int(content[9]))
    x6,y6 =(int(content[10]),int(content[11]))
    x7,y7 =(int(content[12]),int(content[13]))
    x8,y8 =(int(content[14]),int(content[15]))
    x9,y9 =(int(content[16]),int(content[17]))
    box = (710,290,1210,790)
    for i in range(2,maxrow):
        combobox(i)
        time.sleep(0.5)
        chromeMax()
        time.sleep(1)
        pyautogui.click(x1, y1)
        pyautogui.hotkey('ctrl', 'a')
        pyautogui.hotkey('ctrl', 'v')
        pyautogui.press('enter')
        time.sleep(1)
        pyautogui.click(x2, y2)
        pyautogui.hotkey('ctrl', 'a')
        pyautogui.hotkey('ctrl', 'c')
        chromeMin()
        clickin()
        pyautogui.press('right')
        pyautogui.hotkey('ctrl', 'v')
        exlMin()
        chromeMax()
        pyautogui.click(x3, y3)
        pyautogui.hotkey('ctrl', 'a')  
        pyautogui.hotkey('ctrl', 'c')
        chromeMin()
        clickin()
        pyautogui.press('right')
        pyautogui.hotkey('ctrl', 'v')
        pyautogui.press('right')
        pyautogui.hotkey('ctrl', 'c')
        exlMin()
        chromeMax()
        pyautogui.hotkey('ctrl', 'Tab')
        pyautogui.click(x4, y4,clicks=3,duration=0.1)
        pyautogui.hotkey('ctrl', 'v') 
        pyautogui.click(x5, y5)
        time.sleep(1.5)
        pyautogui.click(x6, y6)
        time.sleep(1)
        pyautogui.click(x9, y9,clicks=6)
        time.sleep(0.8)
        p1 = PIL.ImageGrab.grab().crop(box)
        pyautogui.click(x7, y7)
        time.sleep(0.8)
        p2 = PIL.ImageGrab.grab().crop(box)
        pyautogui.click(x7, y7)
        time.sleep(0.8)
        p3 = PIL.ImageGrab.grab().crop(box)
        pyautogui.click(x7, y7)
        time.sleep(0.8)
        p4 = PIL.ImageGrab.grab().crop(box)
        pyautogui.click(x8, y8)
        time.sleep(0.8)
        p5 = PIL.ImageGrab.grab().crop(box)
        if p5 != p1:
            p1.save(path_45%(i,1))
            p2.save(path_45%(i,2))
            p3.save(path_45%(i,3))
            p4.save(path_45%(i,4))
            p5.save(path_top%(i))
            pyautogui.press('esc')
            pyautogui.hotkey('ctrl', 'Tab')
            print(i)
        elif p5 == p1:
            p5.save(path_top%(i))
            pyautogui.press('esc')
            pyautogui.hotkey('ctrl', 'Tab')
            chromeMin()
            print(i)
        exlMin()
    print("Download Complete!")

#initiate the script
#after the first run, script will generate location record txt file for future use
print("**This script requires Google Chrome and 64 bit python 3.6, it runs on python shell only**")
print("**Please save your work before continueing**")
print("**Please save your work before continueing**")
print("**Default directories and Google Static Map API key are listed in the begining of the script, please make sure they are correct before begining.**")
question = input("Ready? y/n: ")
for i in range(1,100):
            if question == 'y':
                app = Application().Connect(title=os.path.basename(__file__) + " - "+ os.path.abspath(__file__) + " (" + python_version()+ ")", class_name='TkTopLevel')
                tktoplevel = app.TkTopLevel
                tktoplevel.Minimize()
                tktoplevel.Close()
                os.startfile(exlpath)
                os.system('taskkill /im chrome.exe')
                time.sleep(3)
                subprocess.Popen([chromepath, chromeurl_coor])
                time.sleep(3)
                subprocess.Popen([chromepath, chromeurl_image])
                time.sleep(3)
                pyautogui.hotkey('ctrl', 'Tab')
                exlMin()
                #configure time depending on website response speed
                time.sleep(10)
                chromeMin()
                shell()
                question = input("First time running this script on this device? y/n: ")
                for l in range(1,100):
                    if question == 'y':
                        configure()
                        recordMouse()
                        break
                    elif question =="n":
                        break
                print("**Mouse and keyboard movement might interrupt downlaod**")
                question = input("Ready ? y/n: ")
                for m in range(1,100):
                    if question == 'y':    
                        break
                generateScript()
                shellMin()
                ImageGrab()      
                        
            

